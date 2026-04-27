#!/usr/bin/env python3
"""
Script to download school images from Google Drive and upload them to Azure
Blob Storage under the structure: schools/{CCT}/{N}.ext

Usage:
    pip install openpyxl gdown azure-storage-blob python-dotenv unidecode
    AZURE_STORAGE_CONNECTION_STRING=<conn_str> python upload_school_images.py
    -- or --
    Set AZURE_STORAGE_CONNECTION_STRING in a .env file next to this script.

The script reads:
  - imgs.txt   : list of school labels and Google Drive folder URLs (in the repo root)
  - escuelas.xlsx : Excel file with school data including CCT codes (in the repo root)
"""

import os
import re
import sys
import logging
import tempfile
import unicodedata
from difflib import SequenceMatcher
from pathlib import Path

import openpyxl
import gdown
from azure.storage.blob import BlobServiceClient
from dotenv import load_dotenv

# ---------------------------------------------------------------------------
# Paths (resolved relative to this script's location, assuming repo structure)
# ---------------------------------------------------------------------------
SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
IMGS_TXT = REPO_ROOT / "imgs.txt"
EXCEL_FILE = REPO_ROOT / "escuelas.xlsx"

AZURE_CONTAINER = "schools"

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def normalize(text: str) -> str:
    """Lowercase, strip accents, collapse whitespace for fuzzy comparison."""
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    return re.sub(r"\s+", " ", text).strip().lower()


def similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, normalize(a), normalize(b)).ratio()


# ---------------------------------------------------------------------------
# Step 1 – Parse Excel: build plantel → [CCT] and escuela → plantel mappings
# ---------------------------------------------------------------------------

def load_school_data(excel_path: Path) -> dict[str, list[str]]:
    """
    Returns a dict mapping plantel name → list[CCT].
    Also handles multiple rows per plantel (one per school level/CCT).
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb["Datos de las escuelas"]

    # Find the header row (the one containing 'Municipio')
    header_row = None
    for row in ws.iter_rows():
        if any(cell.value == "Municipio" for cell in row):
            header_row = [cell.value for cell in row]
            header_row_idx = row[0].row
            break

    if header_row is None:
        raise ValueError("Could not find header row in 'Datos de las escuelas' sheet")

    col = {v: i for i, v in enumerate(header_row)}

    # plantel → set of CCTs (preserve insertion order via dict)
    plantel_ccts: dict[str, list[str]] = {}
    # escuela name → plantel name (for reverse lookup)
    escuela_to_plantel: dict[str, str] = {}

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        plantel = row[col["Plantel"]]
        escuela = row[col["Escuela"]]
        cct = row[col["CCT"]]

        if not plantel or not cct:
            continue

        plantel = str(plantel).strip()
        escuela = str(escuela).strip()
        cct = str(cct).strip()

        if plantel not in plantel_ccts:
            plantel_ccts[plantel] = []
        if cct not in plantel_ccts[plantel]:
            plantel_ccts[plantel].append(cct)

        escuela_to_plantel[escuela] = plantel

    return plantel_ccts, escuela_to_plantel


# ---------------------------------------------------------------------------
# Step 2 – Parse imgs.txt: extract (label, drive_url) pairs
# ---------------------------------------------------------------------------

def parse_imgs_txt(imgs_path: Path) -> list[tuple[str, str]]:
    """Returns list of (school_label, drive_folder_url) pairs."""
    raw = imgs_path.read_bytes()
    # File uses Latin-1 bytes for accented chars and literal '\u2013' for en dashes
    text = raw.decode("latin-1")

    lines = text.splitlines()
    pairs = []
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        if line and not line.startswith("http"):
            label = line
            if i + 1 < len(lines) and lines[i + 1].strip().startswith("http"):
                url = lines[i + 1].strip()
                pairs.append((label, url))
                i += 2
                continue
        i += 1
    return pairs


# ---------------------------------------------------------------------------
# Step 3 – Match drive folder label → plantel → CCTs
# ---------------------------------------------------------------------------

def extract_school_name(label: str) -> str:
    """
    Strips the municipality prefix (e.g. 'ARA \\u2013') and the '(general)' suffix.
    The file stores the en dash as the literal text '\\u2013' (not the Unicode char).
    Example: 'ZAP \\u2013 Carlos de Icaza (general)' → 'Carlos de Icaza'
    """
    # Remove 'XXX \u2013' or 'XXX -' prefix (literal backslash-u-2-0-1-3 or dash)
    label = re.sub(r"^\s*\w+\s*(\\u2013|-)\s*", "", label)
    # Remove trailing parenthetical like '(general)'
    label = re.sub(r"\s*\(.*?\)\s*$", "", label)
    return label.strip()


MIN_MATCH_SCORE = 0.50  # minimum acceptable fuzzy match score


def find_best_plantel(
    school_name: str,
    plantel_ccts: dict[str, list[str]],
    escuela_to_plantel: dict[str, str],
) -> tuple[str | None, float]:
    """
    Returns (best_plantel_name, score).
    Matches against both plantel names and individual escuela names.
    Returns (None, 0.0) if no match exceeds MIN_MATCH_SCORE.
    """
    best_plantel = None
    best_score = 0.0

    # Match directly against plantel names
    for plantel in plantel_ccts:
        score = similarity(school_name, plantel)
        if score > best_score:
            best_score = score
            best_plantel = plantel

    # Match against escuela names (may be more specific, e.g. 'Carlos de Icaza')
    for escuela, plantel in escuela_to_plantel.items():
        score = similarity(school_name, escuela)
        if score > best_score:
            best_score = score
            best_plantel = plantel

    if best_score < MIN_MATCH_SCORE:
        return None, best_score

    return best_plantel, best_score


# ---------------------------------------------------------------------------
# Step 4 – Download Google Drive folder
# ---------------------------------------------------------------------------

def download_drive_folder(url: str, dest_dir: Path) -> list[Path]:
    """
    Downloads all files from a Google Drive shared folder into dest_dir.
    Returns list of downloaded file paths (recursively).
    """
    dest_dir.mkdir(parents=True, exist_ok=True)
    log.info("  Downloading Drive folder → %s", dest_dir)
    gdown.download_folder(url, output=str(dest_dir), quiet=False, use_cookies=False)
    files = sorted(p for p in dest_dir.rglob("*") if p.is_file())
    return files


# ---------------------------------------------------------------------------
# Step 5 – Upload files to Azure Blob Storage
# ---------------------------------------------------------------------------

def upload_files(
    files: list[Path],
    ccts: list[str],
    blob_service: BlobServiceClient,
    container: str,
) -> None:
    """
    Uploads each file to schools/{CCT}/{N}.ext for every CCT in the list.
    N is 1-indexed, based on sorted file order.
    """
    container_client = blob_service.get_container_client(container)

    # Ensure container exists
    try:
        container_client.create_container()
        log.info("  Created container '%s'", container)
    except Exception:
        pass  # Already exists

    for cct in ccts:
        log.info("  Uploading %d file(s) to CCT %s …", len(files), cct)
        for n, file_path in enumerate(files, start=1):
            ext = file_path.suffix  # includes the dot, e.g. '.jpg'
            blob_name = f"{cct}/{n}{ext}"
            blob_client = container_client.get_blob_client(blob_name)

            with open(file_path, "rb") as data:
                blob_client.upload_blob(data, overwrite=True)

            log.info("    ✓ %s → %s/%s", file_path.name, cct, f"{n}{ext}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    # Load .env (if present)
    load_dotenv()

    conn_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
    if not conn_str:
        log.error("AZURE_STORAGE_CONNECTION_STRING is not set. Aborting.")
        sys.exit(1)

    # Load school data
    log.info("Loading school data from %s …", EXCEL_FILE)
    plantel_ccts, escuela_to_plantel = load_school_data(EXCEL_FILE)
    log.info("  Found %d plantels, %d escuelas", len(plantel_ccts), len(escuela_to_plantel))

    # Parse drive links
    log.info("Parsing %s …", IMGS_TXT)
    pairs = parse_imgs_txt(IMGS_TXT)
    log.info("  Found %d drive folders", len(pairs))

    # Build plan: (label, url, school_name, plantel, ccts)
    plan = []
    has_low_confidence = False
    log.info("\n--- Matching drive folders to CCTs ---")
    for label, url in pairs:
        school_name = extract_school_name(label)
        plantel, score = find_best_plantel(school_name, plantel_ccts, escuela_to_plantel)
        ccts = plantel_ccts.get(plantel, []) if plantel else []

        if plantel is None:
            log.error(
                "  %-35s → NO MATCH (best score=%.2f) — skipping",
                school_name, score,
            )
            has_low_confidence = True
        elif score < 0.70:
            log.warning(
                "  %-35s → plantel=%-30s score=%.2f ⚠  CCTs=%s",
                school_name, plantel, score, ccts,
            )
            has_low_confidence = True
        else:
            log.info(
                "  %-35s → plantel=%-30s score=%.2f  CCTs=%s",
                school_name, plantel, score, ccts,
            )
        plan.append((label, url, school_name, plantel, ccts))

    if has_low_confidence:
        log.warning("\nOne or more matches have low confidence (score < 0.70). Review above before continuing.")

    # Confirm before downloading / uploading
    print("\nProceed with download and upload? [y/N] ", end="", flush=True)
    answer = input().strip().lower()
    if answer != "y":
        log.info("Aborted by user.")
        sys.exit(0)

    blob_service = BlobServiceClient.from_connection_string(conn_str)

    with tempfile.TemporaryDirectory(prefix="school_imgs_") as tmp_root:
        tmp_root = Path(tmp_root)

        for label, url, school_name, plantel, ccts in plan:
            if not ccts:
                log.warning("No CCTs found for '%s' (plantel='%s'). Skipping.", school_name, plantel)
                continue

            log.info("\n=== %s ===", label)
            dest_dir = tmp_root / re.sub(r"[^\w]", "_", school_name)

            try:
                files = download_drive_folder(url, dest_dir)
            except Exception as exc:
                log.error("  Failed to download '%s': %s", school_name, exc)
                continue

            if not files:
                log.warning("  No files downloaded for '%s'. Skipping upload.", school_name)
                continue

            log.info("  Downloaded %d file(s)", len(files))

            try:
                upload_files(files, ccts, blob_service, AZURE_CONTAINER)
            except Exception as exc:
                log.error("  Upload failed for '%s': %s", school_name, exc)
                continue

    log.info("\nDone.")


if __name__ == "__main__":
    main()
